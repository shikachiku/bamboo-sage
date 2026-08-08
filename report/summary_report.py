from pathlib import Path
from datetime import datetime
from openpyxl import Workbook

import pandas as pd



# =====================================
# Path
# =====================================

PROJECT_ROOT = Path(__file__).resolve().parent.parent


DATA_ROOT = Path(
    "/mnt/chromeos/shared/GoogleDrive/MyDrive/BambooSage/data"
)


REPORT_ROOT = Path(
    "/mnt/chromeos/shared/GoogleDrive/MyDrive/BambooSage/report"
)


OUTPUT_DIR = REPORT_ROOT / "output"


SYMBOL_FILE = (
    PROJECT_ROOT
    / "invest_symbols.csv"
)



# =====================================
# Timeframe
# =====================================

TIMEFRAMES = {

    "MONTH": "1M.csv",
    "WEEK": "1W.csv",
    "DAY": "1D.csv",

}



# =====================================
# Load symbols
# =====================================

def load_symbols():


    symbols = pd.read_csv(
        SYMBOL_FILE
    )


    symbols = symbols[
        symbols["Enable"] == 1
    ]


    return symbols



# =====================================
# Load latest csv
# =====================================

# =====================================
# Load latest csv
# =====================================

def load_latest_csv(path, cache=None):

    # -----------------------------
    # Cache check
    # -----------------------------

    if cache is not None:

        key = str(path)

        if key in cache:

            return cache[key]


    # -----------------------------
    # File check
    # -----------------------------

    if not path.exists():

        result = (None, None)

        if cache is not None:
            cache[str(path)] = result

        return result



    # -----------------------------
    # Load
    # -----------------------------

    data = pd.read_csv(
        path
    )


    if len(data) == 0:

        result = (None, None)

        if cache is not None:
            cache[str(path)] = result

        return result



    latest = data.iloc[-1]


    if len(data) >= 2:

        previous = data.iloc[-2]

    else:

        previous = None



    result = (
        latest,
        previous
    )


    # -----------------------------
    # Save cache
    # -----------------------------

    if cache is not None:

        cache[str(path)] = result


    return result



# =====================================
# DI direction
# =====================================

def get_di_direction(row):


    if (
        "+DI" not in row.index
        or "-DI" not in row.index
    ):

        return ""


    if row["+DI"] > row["-DI"]:

        return "+"


    return "-"

def get_ha_color(row):


    if row is None:

        return ""


    if "HA_Color" in row.index:

        return row["HA_Color"]


    if "HA_COLOR" in row.index:

        return row["HA_COLOR"]


    return ""
# =====================================
# State
# =====================================

def get_state(
    row,
    previous=None
):


    if row is None:

        return ""


    adx_level = row.get(
        "ADX_LEVEL",
        0
    )


    di = get_di_direction(
        row
    )


    previous_adx = None


    if previous is not None:

        previous_adx = previous.get(
            "ADX_LEVEL",
            None
        )


    # -----------------------------
    # GOLD
    # ADX >= 0.65
    # DI -
    # -----------------------------

    if (

        adx_level >= 0.65

        and

        di == "-"

    ):

        return "GOLD"



    # -----------------------------
    # RED
    # ADX >= 0.65
    # DI +
    # -----------------------------

    if (

        adx_level >= 0.65

        and

        di == "+"

    ):

        return "RED"



    # -----------------------------
    # BLUE
    # ADX < 0.3
    # DI -
    # -----------------------------

    if (

        adx_level < 0.3

        and

        di == "-"

    ):

        return "BLUE"



    # -----------------------------
    # GREEN
    # ADX < 0.3
    # DI +
    # -----------------------------

    if (

        adx_level < 0.3

        and

        di == "+"

    ):

        return "GREEN"



    # -----------------------------
    # Middle zone
    # 0.3 <= ADX < 0.65
    # -----------------------------

    if (

        previous_adx is not None

        and

        0.3 <= adx_level < 0.65

    ):


        # -------------------------
        # ADX rising
        # -------------------------

        if previous_adx < adx_level:


            if di == "+":

                return "YELLOWGREEN"


            if di == "-":

                return "PINK"



        # -------------------------
        # ADX falling
        # -------------------------

        if previous_adx > adx_level:


            if di == "+":

                return "PINK"


            if di == "-":

                return "YELLOWGREEN"



    return ""

# =====================================
# Create symbol data
# =====================================

def create_symbol_row(
    folder,
    name,
    cache
):

    result = {

        "SYMBOL": name

    }


    analysis_dir = (

        DATA_ROOT
        / folder
        / "analysis"

    )


    for tf, filename in TIMEFRAMES.items():


        csv_file = (
            analysis_dir
            / filename
        )


        latest, previous = load_latest_csv(
            csv_file,
            cache
        )


        # -----------------------------
        # No data
        # -----------------------------

        if latest is None:


            result[f"{tf}_PREV_ADX"] = ""
            result[f"{tf}_PREV_DI"] = ""

            result[f"{tf}_ADX"] = ""
            result[f"{tf}_DI"] = ""

            result[f"{tf}_STATE"] = ""
            result[f"{tf}_HA2"] = ""

            result[f"{tf}_CLOSE"] = ""
            result[f"{tf}_HIGH5MA"] = ""
            result[f"{tf}_LOW5MA"] = ""

            continue



        # -----------------------------
        # Previous
        # -----------------------------

        if previous is None:


            result[f"{tf}_PREV_ADX"] = ""
            result[f"{tf}_PREV_DI"] = ""

            prev_ha = ""


        else:


            result[f"{tf}_PREV_ADX"] = previous.get(
                "ADX_LEVEL",
                ""
            )


            result[f"{tf}_PREV_DI"] = get_di_direction(
                previous
            )


            prev_ha = get_ha_color(
                previous
            )



        # -----------------------------
        # Current ADX / DI
        # -----------------------------

        result[f"{tf}_ADX"] = latest.get(
            "ADX_LEVEL",
            ""
        )


        result[f"{tf}_DI"] = get_di_direction(
            latest
        )



        # -----------------------------
        # Current Price
        # -----------------------------

        result[f"{tf}_CLOSE"] = latest.get(
            "Close",
            ""
        )


        result[f"{tf}_HIGH5MA"] = latest.get(
            "High5MA",
            ""
        )


        result[f"{tf}_LOW5MA"] = latest.get(
            "Low5MA",
            ""
        )



        # -----------------------------
        # State
        # -----------------------------

        result[f"{tf}_STATE"] = get_state(
            latest,
            previous
        )



        # -----------------------------
        # HA 2 Pattern
        # -----------------------------

        curr_ha = get_ha_color(
            latest
        )


        if (

            prev_ha == "BLUE"

            and

            curr_ha == "BLUE"

        ):

            result[f"{tf}_HA2"] = "BB"



        elif (

            prev_ha == "BLUE"

            and

            curr_ha == "RED"

        ):

            result[f"{tf}_HA2"] = "BR"



        elif (

            prev_ha == "RED"

            and

            curr_ha == "BLUE"

        ):

            result[f"{tf}_HA2"] = "RB"



        elif (

            prev_ha == "RED"

            and

            curr_ha == "RED"

        ):

            result[f"{tf}_HA2"] = "RR"



        else:

            result[f"{tf}_HA2"] = ""



    return result

# =====================================
# Create report
# =====================================

def create_report():

    symbols = load_symbols()


    rows = []


    # =================================
    # CSV cache
    # =================================

    cache = {}



    dates = {

        "MONTH_DATE": "",
        "WEEK_DATE": "",
        "DAY_DATE": ""

    }



    for _, item in symbols.iterrows():


        analysis_dir = (

            DATA_ROOT
            / item["Folder"]
            / "analysis"

        )



        if dates["MONTH_DATE"] == "":


            for key, filename in TIMEFRAMES.items():


                latest, previous = load_latest_csv(

                    analysis_dir
                    / filename,

                    cache

                )


                if latest is not None:


                    dates[f"{key}_DATE"] = latest.get(

                        "Date",
                        ""

                    )



        rows.append(

            create_symbol_row(

                item["Folder"],
                item["Name"],
                cache

            )

        )



    df = pd.DataFrame(

        rows

    )


    return df, dates



# =====================================
# Save report
# =====================================

def save_report(df, dates):

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True
    )

    REPORT_ROOT.mkdir(
        parents=True,
        exist_ok=True
    )


    # =================================
    # Column order
    # =================================

    column_order = [

        "SYMBOL",

        "MONTH_PREV_ADX",
        "MONTH_PREV_DI",
        "MONTH_ADX",
        "MONTH_DI",
        "MONTH_STATE",
        "MONTH_HA2",

        "MONTH_CLOSE",
        "MONTH_HIGH5MA",
        "MONTH_LOW5MA",


        "WEEK_PREV_ADX",
        "WEEK_PREV_DI",
        "WEEK_ADX",
        "WEEK_DI",
        "WEEK_STATE",
        "WEEK_HA2",

        "WEEK_CLOSE",
        "WEEK_HIGH5MA",
        "WEEK_LOW5MA",


        "DAY_PREV_ADX",
        "DAY_PREV_DI",
        "DAY_ADX",
        "DAY_DI",
        "DAY_STATE",
        "DAY_HA2",

        "DAY_CLOSE",
        "DAY_HIGH5MA",
        "DAY_LOW5MA",

    ]


    df = df.reindex(
        columns=column_order
    )


    # =================================
    # ADX display format
    # =================================

    adx_columns = [

        "MONTH_PREV_ADX",
        "MONTH_ADX",

        "WEEK_PREV_ADX",
        "WEEK_ADX",

        "DAY_PREV_ADX",
        "DAY_ADX",

    ]


    for col in adx_columns:

        if col not in df.columns:

            continue


        df[col] = df[col].apply(

            lambda x:

            f"{float(x):.2f}"

            if x != ""

            else ""

        )


    # =================================
    # Add blank column
    # =================================

    report_df = df.copy()


    report_df.insert(
        0,
        "",
        ""
    )


    # =================================
    # File name
    # =================================

    filename = datetime.now().strftime(
        "%Y-%m-%d_%H-%M.csv"
    )


    output_file = (
        OUTPUT_DIR
        / filename
    )


    latest_file = (
        REPORT_ROOT
        / "latest.csv"
    )


    report_time = datetime.now().strftime(
        "%Y-%m-%d_%H-%M"
    )


    # =================================
    # Save CSV
    # =================================

    with open(
        output_file,
        "w",
        encoding="utf-8-sig",
        newline=""
    ) as f:

        f.write(
            f"REPORT_DATE,{report_time}\n"
        )


        f.write(
            f"MONTH_DATE,{dates['MONTH_DATE']}\n"
        )


        f.write(
            f"WEEK_DATE,{dates['WEEK_DATE']}\n"
        )


        f.write(
            f"DAY_DATE,{dates['DAY_DATE']}\n"
        )


        f.write(
            "\n"
        )


        report_df.to_csv(
            f,
            index=False
        )


    # =================================
    # Update latest CSV
    # =================================

    with open(
        output_file,
        "r",
        encoding="utf-8-sig"
    ) as src:

        content = src.read()


    with open(
        latest_file,
        "w",
        encoding="utf-8-sig"
    ) as dst:

        dst.write(
            content
        )


    print(
        "REPORT SAVED"
    )


    # =================================
    # Save Excel from template
    # =================================

    from openpyxl import load_workbook
    import shutil


    TEMPLATE_FILE = (
        REPORT_ROOT
        / "template"
        / "summary_template.xlsx"
    )


    excel_file = (
        REPORT_ROOT
        / "latest.xlsx"
    )


    # -----------------------------
    # Template copy
    # -----------------------------

    shutil.copy(
        TEMPLATE_FILE,
        excel_file
    )


    # -----------------------------
    # Open template
    # -----------------------------

    workbook = load_workbook(
        excel_file
    )


    sheet = workbook[
        "REPORT"
    ]


    # -----------------------------
    # Clear old data area
    # -----------------------------

    # 既存の結合セルを解除
    for merged_range in list(
        sheet.merged_cells.ranges
    ):

        if merged_range.min_row >= 7:

            sheet.unmerge_cells(
                str(merged_range)
            )


    # -----------------------------
    # Clear old values
    # -----------------------------

    for row in sheet.iter_rows(
        min_row=7
    ):

        for cell in row:

            cell.value = None


    # -----------------------------
    # Header
    # -----------------------------

    sheet["A1"] = "REPORT_DATE"
    sheet["B1"] = report_time

    sheet["A2"] = "MONTH_DATE"
    sheet["B2"] = dates["MONTH_DATE"]

    sheet["A3"] = "WEEK_DATE"
    sheet["B3"] = dates["WEEK_DATE"]

    sheet["A4"] = "DAY_DATE"
    sheet["B4"] = dates["DAY_DATE"]


    # =================================
    # Table Header
    # =================================

    start_row = 6


    headers = [

        "",
        "SYMBOL",

        "MONTH_ADX",
        "MONTH_DI",
        "MONTH_STATE",
        "MONTH_HA2",
        "MONTH_CLOSE",
        "MONTH_5HMA",

        "WEEK_ADX",
        "WEEK_DI",
        "WEEK_STATE",
        "WEEK_HA2",
        "WEEK_CLOSE",
        "WEEK_5HMA",

        "DAY_ADX",
        "DAY_DI",
        "DAY_STATE",
        "DAY_HA2",
        "DAY_CLOSE",
        "DAY_5HMA",

    ]


    for col, value in enumerate(
        headers,
        1
    ):

        sheet.cell(
            start_row,
            col,
            value
        )


    # =================================
    # Table Data
    # =================================

    excel_row = start_row + 1


    for _, row in report_df.iterrows():

        top_row = excel_row
        bottom_row = excel_row + 1


        # ---------------------------------
        # Blank column
        # ---------------------------------

        sheet.cell(
            top_row,
            1,
            ""
        )

        sheet.cell(
            bottom_row,
            1,
            ""
        )


        # ---------------------------------
        # SYMBOL
        # ---------------------------------

        sheet.cell(
            top_row,
            2,
            row["SYMBOL"]
        )


        # =================================
        # MONTH
        # =================================

        # ADX
        sheet.cell(
            top_row,
            3,
            row["MONTH_PREV_ADX"]
        )

        sheet.cell(
            bottom_row,
            3,
            row["MONTH_ADX"]
        )


        # DI
        sheet.cell(
            top_row,
            4,
            row["MONTH_PREV_DI"]
        )

        sheet.cell(
            bottom_row,
            4,
            row["MONTH_DI"]
        )


        # STATE
        sheet.cell(
            top_row,
            5,
            row["MONTH_STATE"]
        )


        # HA2
        sheet.cell(
            top_row,
            6,
            row["MONTH_HA2"]
        )


        # CLOSE
        sheet.cell(
            top_row,
            7,
            row["MONTH_CLOSE"]
        )


        # 5HMA
        sheet.cell(
            top_row,
            8,
            row["MONTH_HIGH5MA"]
        )

        sheet.cell(
            bottom_row,
            8,
            row["MONTH_LOW5MA"]
        )


        # =================================
        # WEEK
        # =================================

        # ADX
        sheet.cell(
            top_row,
            9,
            row["WEEK_PREV_ADX"]
        )

        sheet.cell(
            bottom_row,
            9,
            row["WEEK_ADX"]
        )


        # DI
        sheet.cell(
            top_row,
            10,
            row["WEEK_PREV_DI"]
        )

        sheet.cell(
            bottom_row,
            10,
            row["WEEK_DI"]
        )


        # STATE
        sheet.cell(
            top_row,
            11,
            row["WEEK_STATE"]
        )


        # HA2
        sheet.cell(
            top_row,
            12,
            row["WEEK_HA2"]
        )


        # CLOSE
        sheet.cell(
            top_row,
            13,
            row["WEEK_CLOSE"]
        )


        # 5HMA
        sheet.cell(
            top_row,
            14,
            row["WEEK_HIGH5MA"]
        )

        sheet.cell(
            bottom_row,
            14,
            row["WEEK_LOW5MA"]
        )


        # =================================
        # DAY
        # =================================

        # ADX
        sheet.cell(
            top_row,
            15,
            row["DAY_PREV_ADX"]
        )

        sheet.cell(
            bottom_row,
            15,
            row["DAY_ADX"]
        )


        # DI
        sheet.cell(
            top_row,
            16,
            row["DAY_PREV_DI"]
        )

        sheet.cell(
            bottom_row,
            16,
            row["DAY_DI"]
        )


        # STATE
        sheet.cell(
            top_row,
            17,
            row["DAY_STATE"]
        )


        # HA2
        sheet.cell(
            top_row,
            18,
            row["DAY_HA2"]
        )


        # CLOSE
        sheet.cell(
            top_row,
            19,
            row["DAY_CLOSE"]
        )


        # 5HMA
        sheet.cell(
            top_row,
            20,
            row["DAY_HIGH5MA"]
        )

        sheet.cell(
            bottom_row,
            20,
            row["DAY_LOW5MA"]
        )


        # =================================
        # Merge cells
        # =================================

        # SYMBOL
        sheet.merge_cells(
            start_row=top_row,
            start_column=2,
            end_row=bottom_row,
            end_column=2
        )


        # MONTH
        sheet.merge_cells(
            start_row=top_row,
            start_column=5,
            end_row=bottom_row,
            end_column=5
        )

        sheet.merge_cells(
            start_row=top_row,
            start_column=6,
            end_row=bottom_row,
            end_column=6
        )

        sheet.merge_cells(
            start_row=top_row,
            start_column=7,
            end_row=bottom_row,
            end_column=7
        )


        # WEEK
        sheet.merge_cells(
            start_row=top_row,
            start_column=11,
            end_row=bottom_row,
            end_column=11
        )

        sheet.merge_cells(
            start_row=top_row,
            start_column=12,
            end_row=bottom_row,
            end_column=12
        )

        sheet.merge_cells(
            start_row=top_row,
            start_column=13,
            end_row=bottom_row,
            end_column=13
        )


        # DAY
        sheet.merge_cells(
            start_row=top_row,
            start_column=17,
            end_row=bottom_row,
            end_column=17
        )

        sheet.merge_cells(
            start_row=top_row,
            start_column=18,
            end_row=bottom_row,
            end_column=18
        )

        sheet.merge_cells(
            start_row=top_row,
            start_column=19,
            end_row=bottom_row,
            end_column=19
        )


        # ---------------------------------
        # Next symbol
        # ---------------------------------

        excel_row += 2


    # =================================
    # Save Excel
    # =================================

    workbook.save(
        excel_file
    )


    print(
        excel_file
    )


# =====================================
# Main
# =====================================

def main():


    df, dates = create_report()


    if len(df) == 0:

        print(
            "NO DATA"
        )

        return


    save_report(
        df,
        dates
    )



if __name__ == "__main__":

    main()